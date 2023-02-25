# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
# Installation command line: python3.11 -m pip install openpyxl
from openpyxl import load_workbook


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
    wb = load_workbook('Test.xlsx')
    sheet = wb.active
    print(sheet["A1"].value)
    print(sheet["B1"].value)
    print(sheet["C1"].value)
    print(sheet["D1"].value)
    print(sheet.cell(row=1, column=1).value)
    print(sheet.cell(row=1, column=2).value)
    print(sheet.cell(row=1, column=3).value)
    print(sheet.cell(row=1, column=4).value)
    print(sheet["A1:D2"])
    for row in sheet["A1:D3"]:
        print([x.value for x in row])